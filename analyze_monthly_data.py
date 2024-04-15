#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

def analyze_monthly_data(file_name, month, years):
    # Загрузка данных из файла Excel
    wb = load_workbook(file_name)
    ws = wb["Лист1"]

    data = pd.DataFrame(ws.iter_rows(values_only=True), columns=["Месяц", "Цена, USD"])
    selected_data = data[data['Месяц'].str.contains('|'.join(years)) & data['Месяц'].str.contains(month)]

    new_ws = wb.create_sheet(title=month)
    new_ws.append(["Year", "Price"])
    for _, row in selected_data.iterrows():
        month_year = row['Месяц'].split(' ')
        new_ws.append([month_year[1], float(row['Цена, USD'])])

    chart = LineChart()
    chart.title = f'Price Trend for {month}'
    chart.x_axis.title = 'Year'
    chart.y_axis.title = 'Price'

    data = Reference(new_ws, min_col=2, min_row=1, max_row=4)
    years = Reference(new_ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(years)

    new_ws.add_chart(chart, "D5")

    wb.save(file_name)

file_name = "Статистика.xlsx"
month = input("Введите название месяца: ")
years = input("Введите года через пробел: ").split()

analyze_monthly_data(file_name, month, years)
